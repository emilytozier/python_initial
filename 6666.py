#задача чтобы изъять из множества excel все нужные ячейки и вывести их в остортированном виде

import openpyxl
from openpyxl import load_workbook
import os
import pathlib
from collections import OrderedDict

path = 'C:/Users/User/Documents/Visual Studio 2017/DjangoWebProject1/DjangoWebProject1/app/rogaikopyta'
#for file in os.listdir(path):
    #wb=load_workbook(os.path.join(path, file), read_only=True)
for k in range(1,1001):
    wb = openpyxl.load_workbook('C:/Users/User/Documents/Visual Studio 2017/DjangoWebProject1/DjangoWebProject1/app/rogaikopyta/{id}.xlsx'.format(id=k))
    sheet=wb['Sheet']
    slo={}

    for file in sheet:
        k=sheet.cell(row=2,column=2).value
        z=sheet.cell(row=2,column=4).value
        slo[k]=z
        g=list(slo)
    for i in sorted(g):
        print(i, slo[k])




